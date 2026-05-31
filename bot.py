import os
import re
import io
import csv
import logging
import asyncio
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import List, Tuple, Optional, Dict

import openpyxl
import phonenumbers

from telegram import Update, InputFile, ReplyKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    ConversationHandler, ContextTypes, filters,
)

# ──────────────────────────────────────────────
# TOKEN — set via environment variable BOT_TOKEN
# ──────────────────────────────────────────────
BOT_TOKEN = os.environ.get("BOT_TOKEN", "YOUR_BOT_TOKEN_HERE")

logging.basicConfig(
    format="%(asctime)s %(levelname)s %(message)s",
    level=logging.INFO,
)
log = logging.getLogger(__name__)

# ──────────────────────────────────────────────
# DEFAULT SETTINGS
# ──────────────────────────────────────────────
DEFAULTS = {
    "default_format": "vcf",
    "per_file":       500,
    "contact_name":   "Contact",
    "start_number":   1,
    "basename":       "contacts",
    "file_start":     1,
    "group_name":     "",
    "group_start":    1,
}

# ──────────────────────────────────────────────
# BUTTON LABELS
# ──────────────────────────────────────────────
B_FA = "📊 File Analysis"
B_FC = "🔄 File Converter"
B_QV = "⚡ Quick VCF"
B_VM = "🛠 VCF Maker"
B_SF = "✂️ Split File"
B_MF = "🔗 Merge Files"
B_FE = "✏️ File Editor"
B_RF = "📝 Rename File"
B_RC = "👤 Rename Contact"
B_ST = "⚙️ Settings"
B_RS = "🔄 Reset"
B_HL = "❓ Help"
B_BACK   = "🔙 Back"
B_CANCEL = "❌ Cancel"
B_DONE   = "✅ Done"

MENU_RE = (
    r"^(📊 File Analysis|🔄 File Converter|⚡ Quick VCF|🛠 VCF Maker|"
    r"✂️ Split File|🔗 Merge Files|✏️ File Editor|"
    r"📝 Rename File|👤 Rename Contact|⚙️ Settings|🔄 Reset|❓ Help)$"
)

# ──────────────────────────────────────────────
# KEYBOARDS
# ──────────────────────────────────────────────
def kb_main():
    return ReplyKeyboardMarkup([
        [B_FA, B_FC], [B_QV, B_VM], [B_SF, B_MF],
        [B_FE, B_RF], [B_RC, B_ST], [B_RS, B_HL],
    ], resize_keyboard=True)

def kb_bc():   return ReplyKeyboardMarkup([[B_BACK, B_CANCEL]], resize_keyboard=True, one_time_keyboard=True)
def kb_dc():   return ReplyKeyboardMarkup([[B_DONE, B_CANCEL]], resize_keyboard=True, one_time_keyboard=True)
def kb_fmt():  return ReplyKeyboardMarkup([["📄 TXT","📇 VCF"],["📊 CSV","📑 XLSX"],[B_BACK,B_CANCEL]], resize_keyboard=True, one_time_keyboard=True)
def kb_yn():   return ReplyKeyboardMarkup([["✅ Yes, Reset", B_CANCEL]], resize_keyboard=True, one_time_keyboard=True)
def kb_mf():   return ReplyKeyboardMarkup([["➕ Add More","✅ Finish"],[B_CANCEL]], resize_keyboard=True, one_time_keyboard=True)
def kb_skip(): return ReplyKeyboardMarkup([["⏭ Skip", B_CANCEL]], resize_keyboard=True, one_time_keyboard=True)
def kb_bsc():  return ReplyKeyboardMarkup([[B_BACK,"⏭ Skip",B_CANCEL]], resize_keyboard=True, one_time_keyboard=True)
def kb_gen():  return ReplyKeyboardMarkup([["✅ Generate", B_CANCEL]], resize_keyboard=True, one_time_keyboard=True)
def kb_rn():   return ReplyKeyboardMarkup([["👥 Rename ALL","👤 Rename SINGLE"],[B_CANCEL]], resize_keyboard=True, one_time_keyboard=True)
def kb_ed():   return ReplyKeyboardMarkup([["◀️ Prev","▶️ Next"],["✏️ Edit","🗑 Remove"],["➕ Add","💾 Save"],[B_BACK]], resize_keyboard=True)

def kb_st(s):
    g = s.get("group_name") or "(none)"
    return ReplyKeyboardMarkup([
        [f"📄 Format: {s['default_format'].upper()}", f"📦 Per File: {s['per_file']}"],
        [f"📁 File Base Name: {s.get('basename','contacts')}", f"📂 File Start: {s.get('file_start',1)}"],
        [f"👤 Contact Name: {s['contact_name']}", f"🔢 Contact Start: {s['start_number']}"],
        [f"🏷 Group Name: {g}", f"🔖 Group Start: {s.get('group_start',1)}"],
        [B_BACK, B_CANCEL],
    ], resize_keyboard=True)

# ──────────────────────────────────────────────
# STATES
# ──────────────────────────────────────────────
FA_UP = 0
FC_UP, FC_FMT = 0, 1
QV_FN, QV_NM, QV_PH, QV_MR = 0, 1, 2, 3
VM_UP, VM_BN, VM_CN, VM_PF, VM_CS, VM_FS, VM_GN, VM_GS, VM_CF = range(9)
SF_UP, SF_CT = 0, 1
MF_UP = 0
FE_UP, FE_VW, FE_ES, FE_EN, FE_EP, FE_RI, FE_AN, FE_AP = range(8)
RF_UP, RF_NM = 0, 1
RC_UP, RC_MD, RC_AL, RC_SS, RC_SN = range(5)
ST_MN, ST_PF, ST_CN, ST_SN, ST_BN, ST_FS, ST_GN, ST_GS = range(8)
RS_CF = 0

# ──────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────
def cfg(ctx):
    if "s" not in ctx.user_data:
        ctx.user_data["s"] = DEFAULTS.copy()
    for k, v in DEFAULTS.items():
        ctx.user_data["s"].setdefault(k, v)
    return ctx.user_data["s"]

def clean_name(n):
    return re.sub(r'[^a-zA-Z0-9_\-]', '', n)

def clean_phone(raw):
    d = re.sub(r'\D', '', raw.strip())
    return ('+' + d) if len(d) >= 7 else ""

# Fast VCF parser (no vobject — works on 100k+ files without freezing)
def parse_vcf(data: bytes) -> List[Tuple[str,str]]:
    out, name, phone, inside = [], "", "", False
    for line in data.decode("utf-8", errors="replace").splitlines():
        u = line.strip().upper()
        if u == "BEGIN:VCARD":
            inside, name, phone = True, "", ""
        elif u == "END:VCARD":
            if inside and phone:
                out.append((name or "Contact", phone))
            inside = False
        elif inside:
            if u.startswith("FN:"):
                name = line[3:].strip()
            elif u.startswith("FN;"):
                c = line.find(":"); name = line[c+1:].strip() if c != -1 else ""
            elif u.startswith("TEL") and not phone:
                c = line.find(":")
                if c != -1:
                    p = clean_phone(line[c+1:])
                    if p: phone = p
    return out

def parse_txt(data: bytes) -> List[Tuple[str,str]]:
    out = []
    for line in data.decode("utf-8", errors="replace").splitlines():
        line = line.strip()
        if not line: continue
        if ',' in line:
            a, b = line.split(',', 1)
            p = clean_phone(b)
            if p: out.append((a.strip() or "Contact", p))
        else:
            p = clean_phone(line)
            if p: out.append(("Contact", p))
    return out

def parse_csv(data: bytes) -> List[Tuple[str,str]]:
    out = []
    for row in csv.reader(io.StringIO(data.decode("utf-8", errors="replace"))):
        if not row: continue
        if len(row) >= 2:
            p = clean_phone(row[1])
            if p: out.append((row[0].strip() or "Contact", p))
        else:
            p = clean_phone(row[0])
            if p: out.append(("Contact", p))
    return out

def parse_xlsx(data: bytes) -> List[Tuple[str,str]]:
    out = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        for row in wb.active.iter_rows(values_only=True):
            if not row: continue
            if len(row) >= 2 and row[1] is not None:
                p = clean_phone(str(row[1]))
                if p: out.append((str(row[0]).strip() if row[0] else "Contact", p))
            elif row[0] is not None:
                p = clean_phone(str(row[0]))
                if p: out.append(("Contact", p))
        wb.close()
    except Exception as e:
        log.error(f"XLSX: {e}")
    return out

def parse_file(data: bytes, fname: str) -> List[Tuple[str,str]]:
    ext = Path(fname).suffix.lower()
    if ext == ".vcf":  return parse_vcf(data)
    if ext == ".txt":  return parse_txt(data)
    if ext == ".csv":  return parse_csv(data)
    if ext in (".xlsx",".xls"): return parse_xlsx(data)
    return parse_txt(data)

def to_vcf(contacts):
    buf = io.StringIO()
    for name, phone in contacts:
        buf.write(f"BEGIN:VCARD\nVERSION:3.0\nFN:{name}\nTEL;TYPE=CELL:{phone}\nEND:VCARD\n\n")
    return buf.getvalue().encode()

def to_txt(contacts):
    return "\n".join(p for _,p in contacts).encode()

def to_csv(contacts):
    out = io.StringIO()
    w = csv.writer(out); w.writerow(["Name","Phone"])
    for n,p in contacts: w.writerow([n,p])
    return out.getvalue().encode()

def to_xlsx(contacts):
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(); ws.append(["Name","Phone"])
    for n,p in contacts: ws.append([n,p])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def to_fmt(contacts, fmt):
    f = fmt.lower()
    if f == "vcf":  return to_vcf(contacts)
    if f == "txt":  return to_txt(contacts)
    if f == "csv":  return to_csv(contacts)
    if f == "xlsx": return to_xlsx(contacts)
    return to_vcf(contacts)

def country(phone):
    try:
        r = phonenumbers.region_code_for_number(
            phonenumbers.parse(phone if phone.startswith('+') else '+'+phone))
        return r or "Unknown"
    except: return "Unknown"

def analyze(contacts):
    seen, clean, dup, junk, cc = set(), 0, 0, 0, {}
    for _, ph in contacts:
        d = re.sub(r'\D','',ph)
        if len(d) < 7: junk += 1; continue
        if d in seen: dup += 1
        else:
            seen.add(d); clean += 1
            if clean <= 10_000:
                c = country(ph); cc[c] = cc.get(c,0)+1
    return {"total":len(contacts),"clean":clean,"dup":dup,"junk":junk,"cc":cc,"sampled":min(clean,10_000)}

def pages(contacts, page, per=10):
    tot = len(contacts)
    tp = max(1,(tot+per-1)//per)
    page = max(0,min(page,tp-1))
    s = page*per
    lines = [f"📋 *Contacts* (Page {page+1}/{tp}):\n"]
    for i,(n,p) in enumerate(contacts[s:s+per],start=s+1):
        lines.append(f"`{i}.` {n} — `{p}`")
    return "\n".join(lines), page, tp

# ──────────────────────────────────────────────
# COMMON
# ──────────────────────────────────────────────
async def menu(upd, ctx, txt="Choose a feature:"):
    await upd.effective_message.reply_text(txt, reply_markup=kb_main())

async def cmd_start(upd: Update, ctx: ContextTypes.DEFAULT_TYPE):
    u = upd.effective_user
    ist = timezone(timedelta(hours=5,minutes=30))
    now = datetime.now(ist)
    await upd.effective_message.reply_text(
        f"👋 Welcome, {u.full_name}!\n\n"
        f"📋 Your Profile:\n"
        f"┌ 👤 Name: {u.full_name}\n"
        f"├ 🆔 ID: {u.id}\n"
        f"├ 📛 Username: @{u.username or 'N/A'}\n"
        f"├ 📅 Date: {now.strftime('%d %B %Y')}\n"
        f"└ 🕐 Time: {now.strftime('%I:%M %p')}\n\n"
        f"🤖 VCF Contact Bot — Choose a feature!",
        reply_markup=kb_main(),
    )
    return ConversationHandler.END

async def cancel(upd: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await menu(upd, ctx, "❌ Cancelled.")
    return ConversationHandler.END

async def redir(upd: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await menu(upd, ctx)
    return ConversationHandler.END

FALLBACKS = [
    CommandHandler("start", cmd_start),
    MessageHandler(filters.Regex(r"^❌ Cancel$"), cancel),
    MessageHandler(filters.Regex(MENU_RE), redir),
]
NAV = ~filters.Regex(r"^❌ Cancel$") & ~filters.Regex(MENU_RE)
TXT = filters.TEXT & ~filters.COMMAND & NAV

# helper to download file bytes
async def dl(bot, file_id) -> bytes:
    f = await bot.get_file(file_id)
    return bytes(await f.download_as_bytearray())

# ══════════════════════════════════════════════
# FEATURE 1 — FILE ANALYSIS
# ══════════════════════════════════════════════
async def fa_start(upd, ctx):
    await upd.message.reply_text("📊 *File Analysis*\n\nSend a file (VCF/TXT/CSV/XLSX):", reply_markup=kb_bc(), parse_mode="Markdown")
    return FA_UP

async def fa_up(upd, ctx):
    if upd.message.text and upd.message.text.strip() == B_BACK:
        await menu(upd, ctx); return ConversationHandler.END
    doc = upd.message.document
    if not doc:
        await upd.message.reply_text("Please send a file.", reply_markup=kb_bc()); return FA_UP
    msg = await upd.message.reply_text("⏳ Analyzing...")
    try:
        data = await dl(ctx.bot, doc.file_id)
        contacts, a = await asyncio.to_thread(lambda: (
            (c := parse_file(data, doc.file_name or "f.vcf")), analyze(c)
        )[-1])
        def _run():
            c = parse_file(data, doc.file_name or "f.vcf"); return c, analyze(c)
        contacts, a = await asyncio.to_thread(_run)
        top = sorted(a["cc"].items(), key=lambda x:-x[1])[:10]
        ctxt = "\n".join(f"  • {c}: {n}" for c,n in top)
        if len(a["cc"]) > 10: ctxt += f"\n  ...and {len(a['cc'])-10} more"
        note = f"\n_⚠️ Country data sampled from first {a['sampled']:,} numbers_" if a["clean"] > a["sampled"] else ""
        res = (
            f"📊 *Analysis — {doc.file_name}*\n\n"
            f"📞 Total: *{a['total']:,}*\n"
            f"✅ Unique: *{a['clean']:,}*\n"
            f"🔁 Duplicates: *{a['dup']:,}*\n"
            f"🗑 Junk: *{a['junk']:,}*\n"
        )
        if ctxt: res += f"\n🌍 *Countries:*\n{ctxt}"
        if note: res += f"\n{note}"
        await msg.edit_text(res, parse_mode="Markdown")
    except Exception as e:
        log.error(e); await msg.edit_text(f"❌ Error: {e}")
    await menu(upd, ctx); return ConversationHandler.END

# ══════════════════════════════════════════════
# FEATURE 2 — FILE CONVERTER
# ══════════════════════════════════════════════
async def fc_start(upd, ctx):
    await upd.message.reply_text("🔄 *File Converter*\n\nSend a file:", reply_markup=kb_bc(), parse_mode="Markdown")
    return FC_UP

async def fc_up(upd, ctx):
    if upd.message.text and upd.message.text.strip() == B_BACK:
        await menu(upd, ctx); return ConversationHandler.END
    doc = upd.message.document
    if not doc:
        await upd.message.reply_text("Please send a file.", reply_markup=kb_bc()); return FC_UP
    msg = await upd.message.reply_text("📥 Reading...")
    try:
        data = await dl(ctx.bot, doc.file_id)
        contacts = await asyncio.to_thread(parse_file, data, doc.file_name or "f.vcf")
        ctx.user_data["fc"] = {"contacts": contacts, "stem": clean_name(Path(doc.file_name or "file").stem) or "converted"}
        await msg.delete()
    except Exception as e:
        await msg.edit_text(f"❌ {e}"); return ConversationHandler.END
    await upd.message.reply_text(f"✅ *{len(contacts):,}* contacts.\n\nChoose output format:", reply_markup=kb_fmt(), parse_mode="Markdown")
    return FC_FMT

async def fc_fmt(upd, ctx):
    t = upd.message.text.strip()
    if t == B_BACK: return await fc_start(upd, ctx)
    m = {"📄 TXT":"txt","📇 VCF":"vcf","📊 CSV":"csv","📑 XLSX":"xlsx"}
    if t not in m:
        await upd.message.reply_text("Use the buttons.", reply_markup=kb_fmt()); return FC_FMT
    fmt = m[t]; fc = ctx.user_data.get("fc", {})
    contacts = fc.get("contacts", []); stem = fc.get("stem", "converted")
    msg = await upd.message.reply_text(f"⏳ Converting {len(contacts):,} contacts...")
    try:
        out = await asyncio.to_thread(to_fmt, contacts, fmt)
        await upd.message.reply_document(InputFile(io.BytesIO(out), filename=f"{stem}.{fmt}"),
            caption=f"✅ {len(contacts):,} contacts → {fmt.upper()}")
        await msg.delete()
    except Exception as e:
        await msg.edit_text(f"❌ {e}"); return ConversationHandler.END
    await menu(upd, ctx); return ConversationHandler.END

# ══════════════════════════════════════════════
# FEATURE 3 — QUICK VCF
# ══════════════════════════════════════════════
async def qv_start(upd, ctx):
    ctx.user_data["qv"] = {"contacts": [], "filename": "contacts"}
    await upd.message.reply_text("⚡ *Quick VCF*\n\nEnter filename (no extension):", reply_markup=kb_bc(), parse_mode="Markdown")
    return QV_FN

async def qv_fn(upd, ctx):
    t = upd.message.text.strip()
    if t in (B_BACK, B_CANCEL): await menu(upd, ctx); return ConversationHandler.END
    ctx.user_data["qv"]["filename"] = clean_name(t) or "contacts"
    await upd.message.reply_text("👤 Enter contact base name (e.g. `Customer`):", reply_markup=kb_bc(), parse_mode="Markdown")
    return QV_NM

async def qv_nm(upd, ctx):
    t = upd.message.text.strip()
    if t == B_BACK: return await qv_start(upd, ctx)
    if t == B_CANCEL: await menu(upd, ctx); return ConversationHandler.END
    ctx.user_data["qv"]["base"] = t or "Contact"
    await upd.message.reply_text("📞 Enter phone number(s) (one per line or comma separated):", reply_markup=kb_bc(), parse_mode="Markdown")
    return QV_PH

async def qv_ph(upd, ctx):
    t = upd.message.text.strip() if upd.message.text else ""
    if t == B_BACK: await upd.message.reply_text("👤 Enter contact base name:", reply_markup=kb_bc()); return QV_NM
    if t == B_CANCEL: await menu(upd, ctx); return ConversationHandler.END
    parts = re.split(r'[\n,;\t]+', t)
    candidates = []
    for p in parts:
        p = p.strip()
        if not p: continue
        sub = re.findall(r'\+?\d[\d\s\-\(\)]{5,}', p)
        candidates.extend(sub) if sub else candidates.append(p)
    base = ctx.user_data["qv"].get("base", "Contact")
    existing = ctx.user_data["qv"]["contacts"]
    same = sum(1 for n,_ in existing if re.match(rf'^{re.escape(base)}\s+\d+$', n))
    start = same + 1
    added, bad, seen = [], 0, {p for _,p in existing}
    for raw in candidates:
        ph = clean_phone(raw)
        if not ph: bad += 1; continue
        if ph in seen: continue
        seen.add(ph); added.append((f"{base} {start+len(added)}", ph))
    if not added:
        await upd.message.reply_text("❌ No valid numbers. Try again:", reply_markup=kb_bc()); return QV_PH
    existing.extend(added)
    ctx.user_data["qv"]["contacts"] = existing
    preview = "\n".join(f"`{i}.` *{n}* — `{p}`" for i,(n,p) in enumerate(added[:10],start=start))
    if len(added)>10: preview += f"\n_...and {len(added)-10} more_"
    msg = f"✅ Added *{len(added):,}*:\n\n{preview}\n\n📊 Total: *{len(existing):,}*"
    if bad: msg += f"\n⚠️ Skipped {bad} invalid"
    msg += "\n\nAdd more?"
    await upd.message.reply_text(msg, reply_markup=kb_mf(), parse_mode="Markdown")
    return QV_MR

async def qv_mr(upd, ctx):
    t = upd.message.text.strip()
    if t == "➕ Add More":
        await upd.message.reply_text("👤 Enter contact base name:", reply_markup=kb_bc(), parse_mode="Markdown"); return QV_NM
    if t == "✅ Finish":
        contacts = ctx.user_data["qv"]["contacts"]; fn = ctx.user_data["qv"]["filename"]
        if not contacts: await upd.message.reply_text("No contacts."); await menu(upd, ctx); return ConversationHandler.END
        out = await asyncio.to_thread(to_vcf, contacts)
        await upd.message.reply_document(InputFile(io.BytesIO(out), filename=f"{fn}.vcf"), caption=f"✅ {len(contacts):,} contacts")
        await menu(upd, ctx); return ConversationHandler.END
    if t == B_CANCEL: await menu(upd, ctx); return ConversationHandler.END
    await upd.message.reply_text("Use the buttons.", reply_markup=kb_mf()); return QV_MR

# ══════════════════════════════════════════════
# FEATURE 4 — VCF MAKER (8-step wizard)
# ══════════════════════════════════════════════
async def vm_start(upd, ctx):
    s = cfg(ctx)
    ctx.user_data["vm"] = {k: s.get(k, v) for k,v in [
        ("basename","contacts"),("contact_name","Contact"),("per_file",500),
        ("contact_start",1),("file_start",1),("group_name",""),("group_start",1)
    ]}
    await upd.message.reply_text(
        "🛠 *VCF Maker* — Step 1/8\n\nUpload file with phone numbers (TXT/CSV/XLSX/VCF):",
        reply_markup=kb_bc(), parse_mode="Markdown")
    return VM_UP

def _nav(t):
    if t == B_CANCEL: return "c"
    if t == B_BACK:   return "b"
    if t == "⏭ Skip": return "s"
    return None

async def vm_up(upd, ctx):
    if upd.message.text and upd.message.text.strip() == B_BACK:
        await menu(upd, ctx); return ConversationHandler.END
    doc = upd.message.document
    if not doc:
        await upd.message.reply_text("Send a file.", reply_markup=kb_bc()); return VM_UP
    msg = await upd.message.reply_text("📥 Reading...")
    try:
        data = await dl(ctx.bot, doc.file_id)
        contacts = await asyncio.to_thread(parse_file, data, doc.file_name or "f.txt")
        phones = [p for _,p in contacts]
        if not phones: await msg.edit_text("❌ No valid numbers."); return VM_UP
        ctx.user_data["vm"]["phones"] = phones
        await msg.edit_text(f"✅ Found *{len(phones):,}* numbers.", parse_mode="Markdown")
    except Exception as e:
        await msg.edit_text(f"❌ {e}"); return VM_UP
    cur = ctx.user_data["vm"]["basename"]
    await upd.message.reply_text(f"Step 2/8 — VCF base name (e.g. `madara` → madara1.vcf):\n_Default: `{cur}` — Skip to keep_", reply_markup=kb_bsc(), parse_mode="Markdown")
    return VM_BN

async def vm_bn(upd, ctx):
    t = upd.message.text.strip(); n = _nav(t)
    if n=="c": await menu(upd, ctx); return ConversationHandler.END
    if n=="b":
        await upd.message.reply_text("Step 1/8 — Upload file:", reply_markup=kb_bc()); return VM_UP
    if n!="s": ctx.user_data["vm"]["basename"] = clean_name(t) or ctx.user_data["vm"]["basename"] or "contacts"
    cur = ctx.user_data["vm"]["contact_name"]
    await upd.message.reply_text(f"Step 3/8 — Contact base name:\n_Default: `{cur}` — Skip to keep_", reply_markup=kb_bsc(), parse_mode="Markdown")
    return VM_CN

async def vm_cn(upd, ctx):
    t = upd.message.text.strip(); n = _nav(t)
    if n=="c": await menu(upd, ctx); return ConversationHandler.END
    if n=="b":
        cur = ctx.user_data["vm"]["basename"]
        await upd.message.reply_text(f"Step 2/8 — VCF base name:\n_Default: `{cur}`_", reply_markup=kb_bsc(), parse_mode="Markdown"); return VM_BN
    if n!="s": ctx.user_data["vm"]["contact_name"] = t or ctx.user_data["vm"]["contact_name"] or "Contact"
    cur = ctx.user_data["vm"]["per_file"]
    await upd.message.reply_text(f"Step 4/8 — Contacts per file:\n_Default: `{cur}` — Skip to keep_", reply_markup=kb_bsc(), parse_mode="Markdown")
    return VM_PF

async def vm_pf(upd, ctx):
    t = upd.message.text.strip(); n = _nav(t)
    if n=="c": await menu(upd, ctx); return ConversationHandler.END
    if n=="b":
        cur = ctx.user_data["vm"]["contact_name"]
        await upd.message.reply_text(f"Step 3/8 — Contact base name:\n_Default: `{cur}`_", reply_markup=kb_bsc(), parse_mode="Markdown"); return VM_CN
    if n!="s":
        try:
            v = int(t)
            if v < 1: raise ValueError
            ctx.user_data["vm"]["per_file"] = v
        except ValueError:
            await upd.message.reply_text("Enter a positive number.", reply_markup=kb_bsc()); return VM_PF
    cur = ctx.user_data["vm"]["contact_start"]
    await upd.message.reply_text(f"Step 5/8 — Contact numbering starts from:\n_Default: `{cur}` — Skip to keep_", reply_markup=kb_bsc(), parse_mode="Markdown")
    return VM_CS

async def vm_cs(upd, ctx):
    t = upd.message.text.strip(); n = _nav(t)
    if n=="c": await menu(upd, ctx); return ConversationHandler.END
    if n=="b":
        cur = ctx.user_data["vm"]["per_file"]
        await upd.message.reply_text(f"Step 4/8 — Contacts per file:\n_Default: `{cur}`_", reply_markup=kb_bsc(), parse_mode="Markdown"); return VM_PF
    if n!="s":
        try: ctx.user_data["vm"]["contact_start"] = int(t)
        except ValueError:
            await upd.message.reply_text("Enter a number.", reply_markup=kb_bsc()); return VM_CS
    cur = ctx.user_data["vm"]["file_start"]
    await upd.message.reply_text(f"Step 6/8 — File numbering starts from:\n_Default: `{cur}` — Skip to keep_", reply_markup=kb_bsc(), parse_mode="Markdown")
    return VM_FS

async def vm_fs(upd, ctx):
    t = upd.message.text.strip(); n = _nav(t)
    if n=="c": await menu(upd, ctx); return ConversationHandler.END
    if n=="b":
        cur = ctx.user_data["vm"]["contact_start"]
        await upd.message.reply_text(f"Step 5/8 — Contact start:\n_Default: `{cur}`_", reply_markup=kb_bsc(), parse_mode="Markdown"); return VM_CS
    if n!="s":
        try: ctx.user_data["vm"]["file_start"] = int(t)
        except ValueError:
            await upd.message.reply_text("Enter a number.", reply_markup=kb_bsc()); return VM_FS
    cur = ctx.user_data["vm"].get("group_name","") or "(none)"
    await upd.message.reply_text(f"Step 7/8 — Group tag name (type `-` to clear):\n_Default: `{cur}` — Skip to keep_", reply_markup=kb_bsc(), parse_mode="Markdown")
    return VM_GN

async def vm_gn(upd, ctx):
    t = upd.message.text.strip(); n = _nav(t)
    if n=="c": await menu(upd, ctx); return ConversationHandler.END
    if n=="b":
        cur = ctx.user_data["vm"]["file_start"]
        await upd.message.reply_text(f"Step 6/8 — File start:\n_Default: `{cur}`_", reply_markup=kb_bsc(), parse_mode="Markdown"); return VM_FS
    if n!="s":
        ctx.user_data["vm"]["group_name"] = "" if t=="-" else t
    if not ctx.user_data["vm"].get("group_name"):
        return await vm_confirm_show(upd, ctx)
    cur = ctx.user_data["vm"]["group_start"]
    await upd.message.reply_text(f"Step 8/8 — Group numbering starts from:\n_Default: `{cur}` — Skip to keep_", reply_markup=kb_bsc(), parse_mode="Markdown")
    return VM_GS

async def vm_gs(upd, ctx):
    t = upd.message.text.strip(); n = _nav(t)
    if n=="c": await menu(upd, ctx); return ConversationHandler.END
    if n=="b":
        cur = ctx.user_data["vm"].get("group_name","") or "(none)"
        await upd.message.reply_text(f"Step 7/8 — Group tag name:\n_Default: `{cur}`_", reply_markup=kb_bsc(), parse_mode="Markdown"); return VM_GN
    if n!="s":
        try: ctx.user_data["vm"]["group_start"] = int(t)
        except ValueError:
            await upd.message.reply_text("Enter a number.", reply_markup=kb_bsc()); return VM_GS
    return await vm_confirm_show(upd, ctx)

async def vm_confirm_show(upd, ctx):
    v = ctx.user_data["vm"]
    tot = len(v["phones"]); tf = (tot + v["per_file"] - 1) // v["per_file"]
    fe = v["file_start"] + tf - 1
    grp = f"{v['group_name']} (from {v['group_start']})" if v.get("group_name") else "(none)"
    await upd.message.reply_text(
        f"✅ *Confirm Settings*\n\n"
        f"📞 Numbers: *{tot:,}*\n"
        f"📁 Files: `{v['basename']}{v['file_start']}.vcf` ... `{v['basename']}{fe}.vcf`\n"
        f"👤 Contact name: `{v['contact_name']}`\n"
        f"🔢 Contact start: `{v['contact_start']}`\n"
        f"📦 Per file: `{v['per_file']}`\n"
        f"📂 File start: `{v['file_start']}`\n"
        f"🏷 Group: `{grp}`\n"
        f"📂 Total files: *{tf:,}*",
        reply_markup=kb_gen(), parse_mode="Markdown")
    return VM_CF

async def vm_cf(upd, ctx):
    t = upd.message.text.strip()
    if t in (B_BACK, B_CANCEL): await menu(upd, ctx); return ConversationHandler.END
    if t != "✅ Generate":
        await upd.message.reply_text("Press ✅ Generate or ❌ Cancel.", reply_markup=kb_gen()); return VM_CF
    v = ctx.user_data["vm"]
    phones = v["phones"]; basename = v["basename"]; cname = v["contact_name"]
    per = v["per_file"]; cstart = v["contact_start"]; fstart = v["file_start"]
    gname = v.get("group_name",""); gstart = v.get("group_start",1)
    tot = len(phones); tf = (tot+per-1)//per
    msg = await upd.message.reply_text(f"⏳ Generating {tf:,} file(s) for {tot:,} contacts...")
    cnum = cstart
    for fi, chunk_s in enumerate(range(0, tot, per)):
        chunk = phones[chunk_s:chunk_s+per]
        fnum = fstart+fi; gnum = gstart+fi
        def _build(chunk=chunk, cnum=cnum, gnum=gnum):
            cs = []
            n = cnum
            for ph in chunk:
                nm = f"{cname} {n} {gname} {gnum}" if gname else f"{cname} {n}"
                cs.append((nm, ph)); n += 1
            return to_vcf(cs)
        data = await asyncio.to_thread(_build)
        await upd.message.reply_document(InputFile(io.BytesIO(data), filename=f"{basename}{fnum}.vcf"),
            caption=f"📦 {basename}{fnum}.vcf: {len(chunk):,} contacts")
        cnum += len(chunk)
        if (fi+1) % 10 == 0 and (fi+1) < tf:
            try: await msg.edit_text(f"⏳ Progress: {fi+1}/{tf} files sent...")
            except: pass
    await msg.edit_text(f"✅ Done! {tf:,} file(s), {tot:,} contacts.")
    await menu(upd, ctx); return ConversationHandler.END

# ══════════════════════════════════════════════
# FEATURE 5 — SPLIT FILE
# ══════════════════════════════════════════════
async def sf_start(upd, ctx):
    await upd.message.reply_text("✂️ *Split File*\n\nSend a file to split:", reply_markup=kb_bc(), parse_mode="Markdown")
    return SF_UP

async def sf_up(upd, ctx):
    if upd.message.text and upd.message.text.strip() == B_BACK:
        await menu(upd, ctx); return ConversationHandler.END
    doc = upd.message.document
    if not doc:
        await upd.message.reply_text("Send a file.", reply_markup=kb_bc()); return SF_UP
    msg = await upd.message.reply_text("📥 Reading...")
    try:
        data = await dl(ctx.bot, doc.file_id)
        contacts = await asyncio.to_thread(parse_file, data, doc.file_name or "f.vcf")
        if not contacts: await msg.edit_text("❌ No contacts found."); return SF_UP
        ctx.user_data["sf"] = {"contacts": contacts, "ext": Path(doc.file_name or "f.vcf").suffix.lower().lstrip("."), "stem": clean_name(Path(doc.file_name or "file").stem) or "file"}
        await msg.edit_text(f"✅ *{len(contacts):,}* contacts found.", parse_mode="Markdown")
    except Exception as e:
        await msg.edit_text(f"❌ {e}"); return SF_UP
    await upd.message.reply_text("How many contacts per split file?", reply_markup=kb_bc())
    return SF_CT

async def sf_ct(upd, ctx):
    t = upd.message.text.strip()
    if t == B_BACK: return await sf_start(upd, ctx)
    try:
        n = int(t)
        if n < 1: raise ValueError
    except ValueError:
        await upd.message.reply_text("Enter a positive number.", reply_markup=kb_bc()); return SF_CT
    sf = ctx.user_data.get("sf", {}); contacts = sf.get("contacts",[]); ext = sf.get("ext","vcf"); stem = sf.get("stem","file")
    tot = len(contacts); tp = (tot+n-1)//n
    msg = await upd.message.reply_text(f"⏳ Splitting into {tp:,} file(s)...")
    for i, s in enumerate(range(0, tot, n)):
        chunk = contacts[s:s+n]
        out = await asyncio.to_thread(to_fmt, chunk, ext)
        await upd.message.reply_document(InputFile(io.BytesIO(out), filename=f"{stem}{i+1}.{ext}"),
            caption=f"📦 Part {i+1}: {len(chunk):,} contacts")
        if (i+1) % 10 == 0 and (i+1) < tp:
            try: await msg.edit_text(f"⏳ Progress: {i+1}/{tp} files...")
            except: pass
    await msg.edit_text(f"✅ Split into {tp:,} file(s).")
    await menu(upd, ctx); return ConversationHandler.END

# ══════════════════════════════════════════════
# FEATURE 6 — MERGE FILES
# ══════════════════════════════════════════════
async def mf_start(upd, ctx):
    ctx.user_data["mf"] = {"contacts":[], "count":0}
    await upd.message.reply_text("🔗 *Merge Files*\n\nUpload files one by one. Press ✅ Done when finished.", reply_markup=kb_dc(), parse_mode="Markdown")
    return MF_UP

async def mf_up(upd, ctx):
    t = upd.message.text.strip() if upd.message.text else ""
    if t == B_DONE:
        cs = ctx.user_data["mf"]["contacts"]
        if not cs: await upd.message.reply_text("❌ No contacts."); await menu(upd, ctx); return ConversationHandler.END
        msg = await upd.message.reply_text(f"⏳ Merging {len(cs):,} contacts...")
        out = await asyncio.to_thread(to_vcf, cs)
        await upd.message.reply_document(InputFile(io.BytesIO(out), filename="merged.vcf"),
            caption=f"✅ Merged {len(cs):,} contacts from {ctx.user_data['mf']['count']} files.")
        await msg.delete(); await menu(upd, ctx); return ConversationHandler.END
    if t in (B_BACK, B_CANCEL): await menu(upd, ctx); return ConversationHandler.END
    doc = upd.message.document
    if not doc:
        await upd.message.reply_text("Send a file or press ✅ Done.", reply_markup=kb_dc()); return MF_UP
    try:
        data = await dl(ctx.bot, doc.file_id)
        nc = await asyncio.to_thread(parse_file, data, doc.file_name or "f.vcf")
        ctx.user_data["mf"]["contacts"].extend(nc); ctx.user_data["mf"]["count"] += 1
        await upd.message.reply_text(
            f"📊 Files: *{ctx.user_data['mf']['count']}* | Contacts: *{len(ctx.user_data['mf']['contacts']):,}*\n\nUpload more or press ✅ Done.",
            reply_markup=kb_dc(), parse_mode="Markdown")
    except Exception as e:
        await upd.message.reply_text(f"❌ {e}", reply_markup=kb_dc())
    return MF_UP

# ══════════════════════════════════════════════
# FEATURE 7 — FILE EDITOR
# ══════════════════════════════════════════════
async def fe_start(upd, ctx):
    ctx.user_data["fe"] = {"contacts":[], "page":0, "ext":"vcf", "stem":"file"}
    await upd.message.reply_text("✏️ *File Editor*\n\nUpload a file to edit:", reply_markup=kb_bc(), parse_mode="Markdown")
    return FE_UP

async def fe_up(upd, ctx):
    if upd.message.text and upd.message.text.strip() == B_BACK:
        await menu(upd, ctx); return ConversationHandler.END
    doc = upd.message.document
    if not doc:
        await upd.message.reply_text("Send a file.", reply_markup=kb_bc()); return FE_UP
    msg = await upd.message.reply_text("📥 Reading...")
    try:
        data = await dl(ctx.bot, doc.file_id)
        contacts = await asyncio.to_thread(parse_file, data, doc.file_name or "f.vcf")
        if not contacts: await msg.edit_text("❌ No contacts."); return FE_UP
        ctx.user_data["fe"] = {
            "contacts": list(contacts), "page": 0,
            "ext": Path(doc.file_name or "f.vcf").suffix.lower().lstrip("."),
            "stem": clean_name(Path(doc.file_name or "file").stem) or "file",
        }
        await msg.delete()
    except Exception as e:
        await msg.edit_text(f"❌ {e}"); return ConversationHandler.END
    return await fe_pg(upd, ctx)

async def fe_pg(upd, ctx):
    fe = ctx.user_data.get("fe",{}); contacts = fe.get("contacts",[]); page = fe.get("page",0)
    txt, page, _ = pages(contacts, page)
    ctx.user_data["fe"]["page"] = page
    await upd.effective_message.reply_text(txt + "\n\n_Navigate, edit, remove, add, or save._", reply_markup=kb_ed(), parse_mode="Markdown")
    return FE_VW

async def fe_vw(upd, ctx):
    t = upd.message.text.strip(); fe = ctx.user_data.get("fe",{}); contacts = fe.get("contacts",[])
    if t == "◀️ Prev":
        ctx.user_data["fe"]["page"] = max(0, fe.get("page",0)-1); return await fe_pg(upd, ctx)
    if t == "▶️ Next":
        tp = max(1,(len(contacts)+9)//10); ctx.user_data["fe"]["page"] = min(tp-1, fe.get("page",0)+1); return await fe_pg(upd, ctx)
    if t == "✏️ Edit":
        await upd.message.reply_text(f"Enter contact number to edit (1–{len(contacts)}):", reply_markup=kb_bc()); return FE_ES
    if t == "🗑 Remove":
        await upd.message.reply_text(f"Enter contact number to remove (1–{len(contacts)}):", reply_markup=kb_bc()); return FE_RI
    if t == "➕ Add":
        await upd.message.reply_text("Enter new contact name:", reply_markup=kb_bc()); return FE_AN
    if t == "💾 Save":
        fe = ctx.user_data.get("fe",{}); contacts = fe.get("contacts",[]); ext = fe.get("ext","vcf"); stem = fe.get("stem","file")
        out = await asyncio.to_thread(to_fmt, contacts, ext)
        await upd.effective_message.reply_document(InputFile(io.BytesIO(out), filename=f"{stem}edited.{ext}"), caption=f"✅ Saved {len(contacts):,} contacts.")
        await menu(upd, ctx); return ConversationHandler.END
    if t == B_BACK: await menu(upd, ctx); return ConversationHandler.END
    await upd.message.reply_text("Use the buttons.", reply_markup=kb_ed()); return FE_VW

async def fe_es(upd, ctx):
    t = upd.message.text.strip()
    if t == B_BACK: return await fe_pg(upd, ctx)
    try:
        i = int(t)-1; cs = ctx.user_data["fe"]["contacts"]
        if not (0 <= i < len(cs)): raise IndexError
        ctx.user_data["fe"]["ei"] = i; n,p = cs[i]
        await upd.message.reply_text(f"Editing: *{n}* — `{p}`\n\nNew name (or `-` to keep):", reply_markup=kb_bc(), parse_mode="Markdown")
        return FE_EN
    except: await upd.message.reply_text("Invalid number.", reply_markup=kb_bc()); return FE_ES

async def fe_en(upd, ctx):
    t = upd.message.text.strip()
    if t == B_BACK:
        tot = len(ctx.user_data["fe"]["contacts"])
        await upd.message.reply_text(f"Enter contact number (1–{tot}):", reply_markup=kb_bc()); return FE_ES
    i = ctx.user_data["fe"].get("ei",0); cs = ctx.user_data["fe"]["contacts"]
    if t != "-": cs[i] = (t, cs[i][1])
    await upd.message.reply_text("New phone (or `-` to keep):", reply_markup=kb_bc()); return FE_EP

async def fe_ep(upd, ctx):
    t = upd.message.text.strip()
    if t == B_BACK:
        i = ctx.user_data["fe"].get("ei",0); n,p = ctx.user_data["fe"]["contacts"][i]
        await upd.message.reply_text(f"New name for *{n}* (or `-`):", reply_markup=kb_bc(), parse_mode="Markdown"); return FE_EN
    i = ctx.user_data["fe"].get("ei",0); cs = ctx.user_data["fe"]["contacts"]
    if t != "-":
        ph = clean_phone(t)
        if not ph: await upd.message.reply_text("❌ Invalid phone.", reply_markup=kb_bc()); return FE_EP
        cs[i] = (cs[i][0], ph)
    n,p = cs[i]; await upd.message.reply_text(f"✅ Updated: *{n}* — `{p}`", parse_mode="Markdown")
    return await fe_pg(upd, ctx)

async def fe_ri(upd, ctx):
    t = upd.message.text.strip()
    if t == B_BACK: return await fe_pg(upd, ctx)
    try:
        i = int(t)-1; cs = ctx.user_data["fe"]["contacts"]
        if not (0 <= i < len(cs)): raise IndexError
        rm = cs.pop(i)
        tp = max(1,(len(cs)+9)//10)
        if ctx.user_data["fe"].get("page",0) >= tp: ctx.user_data["fe"]["page"] = max(0,tp-1)
        await upd.message.reply_text(f"✅ Removed: *{rm[0]}* — `{rm[1]}`", parse_mode="Markdown")
        return await fe_pg(upd, ctx)
    except: await upd.message.reply_text("Invalid number.", reply_markup=kb_bc()); return FE_RI

async def fe_an(upd, ctx):
    t = upd.message.text.strip()
    if t == B_BACK: return await fe_pg(upd, ctx)
    ctx.user_data["fe"]["nn"] = t
    await upd.message.reply_text("Enter phone number:", reply_markup=kb_bc()); return FE_AP

async def fe_ap(upd, ctx):
    t = upd.message.text.strip()
    if t == B_BACK: await upd.message.reply_text("Enter new contact name:", reply_markup=kb_bc()); return FE_AN
    ph = clean_phone(t)
    if not ph: await upd.message.reply_text("❌ Invalid phone.", reply_markup=kb_bc()); return FE_AP
    nm = ctx.user_data["fe"].get("nn","Contact")
    ctx.user_data["fe"]["contacts"].append((nm,ph))
    await upd.message.reply_text(f"✅ Added: *{nm}* — `{ph}`", parse_mode="Markdown")
    return await fe_pg(upd, ctx)

# ══════════════════════════════════════════════
# FEATURE 8 — RENAME FILE
# ══════════════════════════════════════════════
async def rf_start(upd, ctx):
    await upd.message.reply_text("📝 *Rename File*\n\nUpload the file:", reply_markup=kb_bc(), parse_mode="Markdown")
    return RF_UP

async def rf_up(upd, ctx):
    if upd.message.text and upd.message.text.strip() == B_BACK:
        await menu(upd, ctx); return ConversationHandler.END
    doc = upd.message.document
    if not doc: await upd.message.reply_text("Send a file.", reply_markup=kb_bc()); return RF_UP
    msg = await upd.message.reply_text("📥 Reading...")
    try:
        data = await dl(ctx.bot, doc.file_id)
        ctx.user_data["rf"] = {"data": data, "ext": Path(doc.file_name or "file").suffix}
        await msg.delete()
    except Exception as e:
        await msg.edit_text(f"❌ {e}"); return ConversationHandler.END
    await upd.message.reply_text("Enter new filename (no extension):", reply_markup=kb_bc()); return RF_NM

async def rf_nm(upd, ctx):
    t = upd.message.text.strip()
    if t == B_BACK: return await rf_start(upd, ctx)
    nm = clean_name(t) or "renamed"; rf = ctx.user_data.get("rf",{})
    await upd.message.reply_document(InputFile(io.BytesIO(rf.get("data",b"")), filename=f"{nm}{rf.get('ext','.vcf')}"),
        caption=f"✅ Renamed to `{nm}{rf.get('ext','.vcf')}`", parse_mode="Markdown")
    await menu(upd, ctx); return ConversationHandler.END

# ══════════════════════════════════════════════
# FEATURE 9 — RENAME CONTACT
# ══════════════════════════════════════════════
async def rc_start(upd, ctx):
    await upd.message.reply_text("👤 *Rename Contact*\n\nUpload a VCF file:", reply_markup=kb_bc(), parse_mode="Markdown")
    return RC_UP

async def rc_up(upd, ctx):
    if upd.message.text and upd.message.text.strip() == B_BACK:
        await menu(upd, ctx); return ConversationHandler.END
    doc = upd.message.document
    if not doc: await upd.message.reply_text("Send a VCF file.", reply_markup=kb_bc()); return RC_UP
    msg = await upd.message.reply_text("📥 Reading...")
    try:
        data = await dl(ctx.bot, doc.file_id)
        contacts = await asyncio.to_thread(parse_vcf, data)
        if not contacts: await msg.edit_text("❌ No contacts."); return RC_UP
        ctx.user_data["rc"] = {"contacts": list(contacts), "stem": clean_name(Path(doc.file_name or "file").stem) or "file"}
        await msg.delete()
    except Exception as e:
        await msg.edit_text(f"❌ {e}"); return ConversationHandler.END
    await upd.message.reply_text(f"✅ *{len(ctx.user_data['rc']['contacts']):,}* contacts.\n\nRename ALL or SINGLE?", reply_markup=kb_rn(), parse_mode="Markdown")
    return RC_MD

async def rc_md(upd, ctx):
    t = upd.message.text.strip()
    if t in (B_BACK, B_CANCEL): await menu(upd, ctx); return ConversationHandler.END
    if t == "👥 Rename ALL":
        await upd.message.reply_text("Enter new base name (e.g. `Customer`):", reply_markup=kb_bc(), parse_mode="Markdown"); return RC_AL
    if t == "👤 Rename SINGLE":
        cs = ctx.user_data["rc"]["contacts"]
        lines = ["Choose a contact:\n"] + [f"`{i}.` {n} — `{p}`" for i,(n,p) in enumerate(cs[:50],1)]
        if len(cs)>50: lines.append(f"\n_(first 50 of {len(cs):,})_")
        lines.append("\nEnter contact number:")
        await upd.message.reply_text("\n".join(lines), reply_markup=kb_bc(), parse_mode="Markdown"); return RC_SS
    await upd.message.reply_text("Use the buttons.", reply_markup=kb_rn()); return RC_MD

async def rc_al(upd, ctx):
    t = upd.message.text.strip()
    if t == B_BACK: await upd.message.reply_text("Rename ALL or SINGLE?", reply_markup=kb_rn()); return RC_MD
    base = t or "Contact"; cs = ctx.user_data["rc"]["contacts"]
    msg = await upd.message.reply_text(f"⏳ Renaming {len(cs):,} contacts...")
    def _r(): renamed = [(f"{base} {i}",p) for i,(_,p) in enumerate(cs,1)]; return renamed, to_vcf(renamed)
    renamed, out = await asyncio.to_thread(_r)
    ctx.user_data["rc"]["contacts"] = renamed; stem = ctx.user_data["rc"]["stem"]
    await upd.message.reply_document(InputFile(io.BytesIO(out), filename=f"{stem}renamed.vcf"),
        caption=f"✅ Renamed {len(renamed):,} contacts as '{base} N'.")
    await msg.delete(); await menu(upd, ctx); return ConversationHandler.END

async def rc_ss(upd, ctx):
    t = upd.message.text.strip()
    if t == B_BACK: await upd.message.reply_text("Rename ALL or SINGLE?", reply_markup=kb_rn()); return RC_MD
    try:
        i = int(t)-1; cs = ctx.user_data["rc"]["contacts"]
        if not (0 <= i < len(cs)): raise IndexError
        ctx.user_data["rc"]["ei"] = i; n,p = cs[i]
        await upd.message.reply_text(f"Renaming: *{n}* — `{p}`\n\nEnter new name:", reply_markup=kb_bc(), parse_mode="Markdown")
        return RC_SN
    except: await upd.message.reply_text("Invalid number.", reply_markup=kb_bc()); return RC_SS

async def rc_sn(upd, ctx):
    t = upd.message.text.strip()
    if t == B_BACK:
        cs = ctx.user_data["rc"]["contacts"]
        lines = ["Choose a contact:\n"] + [f"`{i}.` {n} — `{p}`" for i,(n,p) in enumerate(cs[:50],1)]
        lines.append("\nEnter contact number:")
        await upd.message.reply_text("\n".join(lines), reply_markup=kb_bc(), parse_mode="Markdown"); return RC_SS
    i = ctx.user_data["rc"].get("ei",0); cs = ctx.user_data["rc"]["contacts"]
    cs[i] = (t, cs[i][1]); msg = await upd.message.reply_text("⏳ Saving...")
    stem = ctx.user_data["rc"]["stem"]
    out = await asyncio.to_thread(to_vcf, cs)
    await upd.message.reply_document(InputFile(io.BytesIO(out), filename=f"{stem}renamed.vcf"), caption=f"✅ Renamed to '{t}'.")
    await msg.delete(); await menu(upd, ctx); return ConversationHandler.END

# ══════════════════════════════════════════════
# FEATURE 10 — SETTINGS
# ══════════════════════════════════════════════
async def st_start(upd, ctx):
    s = cfg(ctx)
    await upd.message.reply_text("⚙️ *Settings*\n\nTap a setting to change it:", reply_markup=kb_st(s), parse_mode="Markdown")
    return ST_MN

async def _st_menu(upd, ctx):
    s = cfg(ctx)
    await upd.message.reply_text("⚙️ *Settings*", reply_markup=kb_st(s), parse_mode="Markdown")
    return ST_MN

async def st_mn(upd, ctx):
    t = upd.message.text.strip(); s = cfg(ctx)
    if t in (B_BACK, B_CANCEL): await menu(upd, ctx); return ConversationHandler.END
    if t.startswith("📄 Format:"):
        fmts = ["vcf","txt","csv","xlsx"]; cur = s["default_format"]
        s["default_format"] = fmts[(fmts.index(cur)+1) % len(fmts)] if cur in fmts else "vcf"
        await upd.message.reply_text(f"✅ Format → *{s['default_format'].upper()}*", reply_markup=kb_st(s), parse_mode="Markdown"); return ST_MN
    if t.startswith("📦 Per File:"):
        await upd.message.reply_text(f"Current: *{s['per_file']}*\n\nEnter new count:", reply_markup=kb_bc(), parse_mode="Markdown"); return ST_PF
    if t.startswith("👤 Contact Name:"):
        await upd.message.reply_text(f"Current: *{s['contact_name']}*\n\nEnter new name:", reply_markup=kb_bc(), parse_mode="Markdown"); return ST_CN
    if t.startswith("🔢 Contact Start:") or t.startswith("🔢 Start Number:"):
        await upd.message.reply_text(f"Current: *{s['start_number']}*\n\nEnter new start:", reply_markup=kb_bc(), parse_mode="Markdown"); return ST_SN
    if t.startswith("📁 File Base Name:"):
        await upd.message.reply_text(f"Current: *{s.get('basename','contacts')}*\n\nEnter new base name:", reply_markup=kb_bc(), parse_mode="Markdown"); return ST_BN
    if t.startswith("📂 File Start:"):
        await upd.message.reply_text(f"Current: *{s.get('file_start',1)}*\n\nEnter new file start:", reply_markup=kb_bc(), parse_mode="Markdown"); return ST_FS
    if t.startswith("🏷 Group Name:"):
        await upd.message.reply_text(f"Current: *{s.get('group_name','') or '(none)'}*\n\nEnter group name (Skip to clear):", reply_markup=kb_skip(), parse_mode="Markdown"); return ST_GN
    if t.startswith("🔖 Group Start:"):
        await upd.message.reply_text(f"Current: *{s.get('group_start',1)}*\n\nEnter group start:", reply_markup=kb_bc(), parse_mode="Markdown"); return ST_GS
    await upd.message.reply_text("Use the buttons.", reply_markup=kb_st(s)); return ST_MN

async def _st_bc(upd, ctx, t):
    if t == B_CANCEL: await menu(upd, ctx); return ConversationHandler.END
    if t == B_BACK: return await _st_menu(upd, ctx)
    return None

async def st_pf(upd, ctx):
    t = upd.message.text.strip(); s = cfg(ctx); r = await _st_bc(upd, ctx, t)
    if r is not None: return r
    try:
        v = int(t)
        if v < 1: raise ValueError
        s["per_file"] = v; await upd.message.reply_text(f"✅ Per file → *{v}*", reply_markup=kb_st(s), parse_mode="Markdown")
    except: await upd.message.reply_text("Enter a positive number.", reply_markup=kb_bc()); return ST_PF
    return ST_MN

async def st_cn(upd, ctx):
    t = upd.message.text.strip(); s = cfg(ctx); r = await _st_bc(upd, ctx, t)
    if r is not None: return r
    if not t: await upd.message.reply_text("Name can't be empty.", reply_markup=kb_bc()); return ST_CN
    s["contact_name"] = t; await upd.message.reply_text(f"✅ Contact name → *{t}*", reply_markup=kb_st(s), parse_mode="Markdown")
    return ST_MN

async def st_sn(upd, ctx):
    t = upd.message.text.strip(); s = cfg(ctx); r = await _st_bc(upd, ctx, t)
    if r is not None: return r
    try: s["start_number"] = int(t); await upd.message.reply_text(f"✅ Contact start → *{t}*", reply_markup=kb_st(s), parse_mode="Markdown")
    except: await upd.message.reply_text("Enter a number.", reply_markup=kb_bc()); return ST_SN
    return ST_MN

async def st_bn(upd, ctx):
    t = upd.message.text.strip(); s = cfg(ctx); r = await _st_bc(upd, ctx, t)
    if r is not None: return r
    n = clean_name(t) or "contacts"; s["basename"] = n
    await upd.message.reply_text(f"✅ Base name → *{n}*", reply_markup=kb_st(s), parse_mode="Markdown")
    return ST_MN

async def st_fs(upd, ctx):
    t = upd.message.text.strip(); s = cfg(ctx); r = await _st_bc(upd, ctx, t)
    if r is not None: return r
    try: s["file_start"] = int(t); await upd.message.reply_text(f"✅ File start → *{t}*", reply_markup=kb_st(s), parse_mode="Markdown")
    except: await upd.message.reply_text("Enter a number.", reply_markup=kb_bc()); return ST_FS
    return ST_MN

async def st_gn(upd, ctx):
    t = upd.message.text.strip(); s = cfg(ctx)
    if t == B_CANCEL: await menu(upd, ctx); return ConversationHandler.END
    if t == B_BACK: return await _st_menu(upd, ctx)
    if t == "⏭ Skip": s["group_name"] = ""; await upd.message.reply_text("✅ Group name cleared.", reply_markup=kb_st(s)); return ST_MN
    s["group_name"] = t; await upd.message.reply_text(f"✅ Group name → *{t}*", reply_markup=kb_st(s), parse_mode="Markdown")
    return ST_MN

async def st_gs(upd, ctx):
    t = upd.message.text.strip(); s = cfg(ctx); r = await _st_bc(upd, ctx, t)
    if r is not None: return r
    try: s["group_start"] = int(t); await upd.message.reply_text(f"✅ Group start → *{t}*", reply_markup=kb_st(s), parse_mode="Markdown")
    except: await upd.message.reply_text("Enter a number.", reply_markup=kb_bc()); return ST_GS
    return ST_MN

# ══════════════════════════════════════════════
# FEATURE 11 — RESET
# ══════════════════════════════════════════════
async def rs_start(upd, ctx):
    await upd.message.reply_text("🔄 *Reset* — Reset all settings to defaults?", reply_markup=kb_yn(), parse_mode="Markdown")
    return RS_CF

async def rs_cf(upd, ctx):
    if upd.message.text.strip() == "✅ Yes, Reset":
        ctx.user_data["s"] = DEFAULTS.copy()
        await upd.message.reply_text("✅ Settings reset.")
    await menu(upd, ctx); return ConversationHandler.END

# ══════════════════════════════════════════════
# FEATURE 12 — HELP
# ══════════════════════════════════════════════
async def hl_start(upd, ctx):
    await upd.message.reply_text(
        "❓ *VCF Contact Bot — Help*\n\n"
        "*📊 File Analysis* — Upload file → get stats + country breakdown\n"
        "*🔄 File Converter* — Convert VCF/TXT/CSV/XLSX to any format\n"
        "*⚡ Quick VCF* — Create VCF from scratch by entering numbers\n"
        "*🛠 VCF Maker* — 8-step wizard: upload numbers → generate VCF files\n"
        "*✂️ Split File* — Split a file into smaller parts\n"
        "*🔗 Merge Files* — Merge multiple files into one VCF\n"
        "*✏️ File Editor* — Browse, edit, remove, add contacts\n"
        "*📝 Rename File* — Rename any file\n"
        "*👤 Rename Contact* — Rename all or single contacts in a VCF\n"
        "*⚙️ Settings* — Set defaults for VCF Maker\n"
        "*🔄 Reset* — Reset settings to defaults\n\n"
        "📞 *Accepted phone formats:*\n"
        "+919876543210, 919876543210, 9876543210\n"
        "_Any format with 7+ digits. All numbers get + prefix._\n\n"
        "📁 *Filenames:* letters, digits, `-`, `_` only",
        reply_markup=kb_main(), parse_mode="Markdown")
    return ConversationHandler.END

# ──────────────────────────────────────────────
# CONVERSATION HANDLER BUILDERS
# ──────────────────────────────────────────────
def conv(btn, states, entry):
    return ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{re.escape(btn)}$"), entry)],
        states=states, fallbacks=FALLBACKS, allow_reentry=True, per_message=False,
    )

def build_handlers():
    D = filters.Document.ALL
    return [
        conv(B_FA, {FA_UP: [MessageHandler(D|TXT, fa_up)]}, fa_start),
        conv(B_FC, {FC_UP: [MessageHandler(D|TXT, fc_up)], FC_FMT: [MessageHandler(TXT, fc_fmt)]}, fc_start),
        conv(B_QV, {QV_FN:[MessageHandler(TXT,qv_fn)],QV_NM:[MessageHandler(TXT,qv_nm)],QV_PH:[MessageHandler(TXT,qv_ph)],QV_MR:[MessageHandler(TXT,qv_mr)]}, qv_start),
        conv(B_VM, {
            VM_UP:[MessageHandler(D|TXT,vm_up)], VM_BN:[MessageHandler(TXT,vm_bn)],
            VM_CN:[MessageHandler(TXT,vm_cn)],   VM_PF:[MessageHandler(TXT,vm_pf)],
            VM_CS:[MessageHandler(TXT,vm_cs)],   VM_FS:[MessageHandler(TXT,vm_fs)],
            VM_GN:[MessageHandler(TXT,vm_gn)],   VM_GS:[MessageHandler(TXT,vm_gs)],
            VM_CF:[MessageHandler(TXT,vm_cf)],
        }, vm_start),
        conv(B_SF, {SF_UP:[MessageHandler(D|TXT,sf_up)], SF_CT:[MessageHandler(TXT,sf_ct)]}, sf_start),
        conv(B_MF, {MF_UP:[MessageHandler(D|TXT,mf_up)]}, mf_start),
        conv(B_FE, {
            FE_UP:[MessageHandler(D|TXT,fe_up)], FE_VW:[MessageHandler(TXT,fe_vw)],
            FE_ES:[MessageHandler(TXT,fe_es)],   FE_EN:[MessageHandler(TXT,fe_en)],
            FE_EP:[MessageHandler(TXT,fe_ep)],   FE_RI:[MessageHandler(TXT,fe_ri)],
            FE_AN:[MessageHandler(TXT,fe_an)],   FE_AP:[MessageHandler(TXT,fe_ap)],
        }, fe_start),
        conv(B_RF, {RF_UP:[MessageHandler(D|TXT,rf_up)], RF_NM:[MessageHandler(TXT,rf_nm)]}, rf_start),
        conv(B_RC, {
            RC_UP:[MessageHandler(D|TXT,rc_up)], RC_MD:[MessageHandler(TXT,rc_md)],
            RC_AL:[MessageHandler(TXT,rc_al)],   RC_SS:[MessageHandler(TXT,rc_ss)],
            RC_SN:[MessageHandler(TXT,rc_sn)],
        }, rc_start),
        conv(B_ST, {
            ST_MN:[MessageHandler(TXT,st_mn)], ST_PF:[MessageHandler(TXT,st_pf)],
            ST_CN:[MessageHandler(TXT,st_cn)], ST_SN:[MessageHandler(TXT,st_sn)],
            ST_BN:[MessageHandler(TXT,st_bn)], ST_FS:[MessageHandler(TXT,st_fs)],
            ST_GN:[MessageHandler(TXT,st_gn)], ST_GS:[MessageHandler(TXT,st_gs)],
        }, st_start),
        conv(B_RS, {RS_CF:[MessageHandler(TXT,rs_cf)]}, rs_start),
        conv(B_HL, {}, hl_start),
    ]

# ──────────────────────────────────────────────
# ENTRY POINT
# ──────────────────────────────────────────────
def main():
    if BOT_TOKEN == "YOUR_BOT_TOKEN_HERE":
        raise SystemExit("Set BOT_TOKEN environment variable before running.\n"
                         "Example: export BOT_TOKEN=123456:ABC-your-token")
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", cmd_start))
    for h in build_handlers():
        app.add_handler(h)
    log.info("Bot started (polling) — no web server, no port binding.")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    main()
